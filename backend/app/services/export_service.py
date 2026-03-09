"""
数据导出服务
"""
from sqlalchemy.orm import Session
from app.models.form import Submission, FormVersion
from app.schemas.submission_schemas import ExportRequest
from app.core.exceptions import BusinessError, NotFoundError
from app.core.redis_client import redis_client
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from pathlib import Path
import uuid
import json
import re
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """导出服务"""

    @staticmethod
    def export_submissions(
            request: ExportRequest,
            tenant_id: int,
            user_id: int,
            db: Session
    ) -> Dict[str, Any]:
        """导出提交数据"""
        # 查询数据量
        query = db.query(Submission).filter(
            Submission.form_id == request.form_id,
            Submission.tenant_id == tenant_id
        )

        if request.submission_ids:
            query = query.filter(Submission.id.in_(request.submission_ids))

        total = query.count()

        # 判断同步还是异步
        from app.config import settings
        threshold = settings.EXPORT_ASYNC_THRESHOLD

        if total <= threshold:
            # 同步导出
            submissions = query.all()
            version = db.query(FormVersion).filter(
                FormVersion.form_id == request.form_id,
                FormVersion.version > 0
            ).order_by(FormVersion.version.desc()).first()

            if not version:
                raise NotFoundError("表单版本不存在")

            # 生成文件
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"export_{request.form_id}_{timestamp}.xlsx"
            output_path = Path(settings.EXPORT_TEMP_DIR) / filename
            output_path.parent.mkdir(parents=True, exist_ok=True)

            ExportService.export_to_excel(
                submissions=submissions,
                form_version=version,
                field_ids=request.field_ids,
                desensitize=request.desensitize,
                output_path=str(output_path)
            )

            # 返回下载URL
            from app.services.storage_service import StorageService
            # 上传到MinIO
            with open(output_path, 'rb') as f:
                upload_result = StorageService.upload_file(
                    file_data=f.read(),
                    filename=filename,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    tenant_id=tenant_id,
                    category="exports"
                )

            download_url = StorageService.get_download_url(upload_result["storage_path"], expires=86400)

            return {
                "download_url": download_url,
                "total_rows": total
            }
        else:
            # 异步导出
            task_id = ExportService.create_export_task(request, tenant_id, user_id)

            # TODO: 使用Celery或后台任务执行
            # 这里简化处理，实际应该用异步任务队列

            return {
                "task_id": task_id,
                "total_rows": total
            }

    @staticmethod
    def export_to_excel(
            submissions: List[Submission],
            form_version: FormVersion,
            field_ids: Optional[List[str]],
            desensitize: bool,
            output_path: str
    ) -> str:
        """导出为Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "提交数据"

        # 获取字段信息
        schema_json = form_version.schema_json
        fields = schema_json.get("fields", [])

        # 确定导出字段
        if field_ids:
            export_fields = [f for f in fields if f.get("id") in field_ids]
        else:
            export_fields = [f for f in fields if f.get("type") != "description"]

        # 表头
        headers = ["提交ID", "提交时间", "提交人", "状态"]
        headers.extend([f.get("label", f.get("id")) for f in export_fields])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, submission in enumerate(submissions, 2):
            data = submission.data_jsonb

            # 脱敏处理
            if desensitize:
                data = ExportService.desensitize_data(data, schema_json)

            # 基础列
            ws.cell(row=row_idx, column=1, value=submission.id)
            ws.cell(row=row_idx, column=2, value=submission.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_idx, column=3, value=submission.submitter_user_id or "匿名")
            ws.cell(row=row_idx, column=4, value=submission.status)

            # 字段数据
            for col_idx, field in enumerate(export_fields, 5):
                field_id = field.get("id")
                value = data.get(field_id)

                # 格式化值
                formatted_value = ExportService._format_field_value(value, field, submission.snapshot_json)
                ws.cell(row=row_idx, column=col_idx, value=formatted_value)

        # 保存文件
        wb.save(output_path)
        logger.info(f"Exported to Excel: {output_path}, rows={len(submissions)}")

        return output_path

    @staticmethod
    def desensitize_data(
            data: Dict[str, Any],
            schema_json: Dict[str, Any]
    ) -> Dict[str, Any]:
        """数据脱敏"""
        result = data.copy()
        fields = schema_json.get("fields", [])

        for field in fields:
            field_id = field.get("id")
            field_type = field.get("type")
            field_label = field.get("label", "").lower()
            value = result.get(field_id)

            if not value or not isinstance(value, str):
                continue

            # 手机号脱敏
            if field_type == "phone" or "phone" in field_label or "手机" in field_label:
                if len(value) == 11:
                    result[field_id] = value[:3] + "****" + value[7:]

            # 邮箱脱敏
            elif field_type == "email" or "email" in field_label or "邮箱" in field_label:
                if "@" in value:
                    parts = value.split("@")
                    if len(parts[0]) > 2:
                        result[field_id] = parts[0][:2] + "***@" + parts[1]

            # 身份证脱敏
            elif "idcard" in field_label or "身份证" in field_label:
                if len(value) >= 15:
                    result[field_id] = value[:3] + "***********" + value[-4:]

        return result

    @staticmethod
    def create_export_task(
            request: ExportRequest,
            tenant_id: int,
            user_id: int
    ) -> str:
        """创建异步导出任务"""
        task_id = str(uuid.uuid4())

        # 保存任务信息到Redis
        task_data = {
            "task_id": task_id,
            "status": "pending",
            "progress": 0,
            "form_id": request.form_id,
            "tenant_id": tenant_id,
            "user_id": user_id,
            "request": request.dict(),
            "created_at": datetime.now().isoformat()
        }

        key = redis_client.get_key(tenant_id, "export", f"task:{task_id}")
        redis_client.setex(key, 86400, json.dumps(task_data))  # 24小时过期

        logger.info(f"Created export task: {task_id}")
        return task_id

    @staticmethod
    def get_export_task_status(
            task_id: str,
            tenant_id: int
    ) -> Dict[str, Any]:
        """查询导出任务状态"""
        key = redis_client.get_key(tenant_id, "export", f"task:{task_id}")
        data = redis_client.get(key)

        if not data:
            raise NotFoundError("导出任务不存在或已过期")

        task_data = json.loads(data)

        return {
            "task_id": task_id,
            "status": task_data.get("status"),
            "progress": task_data.get("progress", 0),
            "download_url": task_data.get("download_url"),
            "error_message": task_data.get("error_message"),
            "created_at": task_data.get("created_at"),
            "expires_at": (datetime.fromisoformat(task_data.get("created_at")) + timedelta(hours=24)).isoformat()
        }

    @staticmethod
    def _format_field_value(
            value: Any,
            field: Dict[str, Any],
            snapshot: Dict[str, Any]
    ) -> str:
        """格式化字段值"""
        if value is None:
            return ""

        field_type = field.get("type")
        field_id = field.get("id")

        # 选项类字段：显示标签
        if field_type in ["select", "radio", "checkbox"]:
            options = snapshot.get("field_options", {}).get(field_id, [])
            if isinstance(value, list):
                labels = []
                for v in value:
                    opt = next((o for o in options if o.get("value") == v), None)
                    labels.append(opt.get("label", v) if opt else str(v))
                return ", ".join(labels)
            else:
                opt = next((o for o in options if o.get("value") == value), None)
                return opt.get("label", value) if opt else str(value)

        # 日期范围
        elif isinstance(value, dict) and "start" in value and "end" in value:
            return f"{value.get('start')} ~ {value.get('end')}"

        # 附件
        elif field_type in ["upload", "image"] and isinstance(value, list):
            return f"{len(value)} 个文件"

        # 数组
        elif isinstance(value, list):
            return ", ".join(str(v) for v in value)

        # 其他
        else:
            return str(value)