"""
模块用途: 批量用户导入服务
依赖配置: openpyxl (Excel解析), passlib (密码加密)
数据流向: Excel文件 -> 解析验证 -> 创建用户 -> 关联部门岗位 -> 返回结果
函数清单:
    - BatchImportService.generate_template(): 生成Excel模板
    - BatchImportService.parse_excel(): 解析Excel文件
    - BatchImportService.validate_row(): 验证单行数据
    - BatchImportService.import_users(): 批量导入用户
    - BatchImportService.save_import_log(): 保存导入日志
    - BatchImportService.preview_import(): 预览导入数据
    - BatchImportService.confirm_import(): 确认导入数据
"""
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from io import BytesIO
import logging
import json
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.user import (
    User, Department, Position, UserProfile,
    UserDepartment, UserDepartmentPost, UserRole, Role
)
from app.models.batch_import import BatchImportLog
from app.core.security import hash_password
from app.core.exceptions import ValidationError, BusinessError
from app.core.redis_client import redis_client
from app.schemas.batch_import import (
    BatchImportUserRow, BatchImportRowResult, BatchImportResponse
)
from app.schemas.user_management import (
    ImportPreviewRow, ImportPreviewResponse
)

logger = logging.getLogger(__name__)


class BatchImportService:
    """批量导入服务"""

    # Excel列映射
    COLUMN_MAP = {
        'A': 'account',
        'B': 'name',
        'C': 'identity_type',
        'D': 'department_name',
        'E': 'position_name',
        'F': 'email',
        'G': 'phone',
        'H': 'identity_no',
        'I': 'entry_year',
        'J': 'grade',
        'K': 'major',
        'L': 'title',
        'M': 'research_area',
        'N': 'office',
        'O': 'emergency_contact',
        'P': 'emergency_phone'
    }

    # 必填列
    REQUIRED_COLUMNS = ['A', 'B', 'C', 'D']

    @staticmethod
    def generate_template(tenant_id: int, db: Session) -> BytesIO:
        """
        生成Excel模板

        Args:
            tenant_id: 租户ID
            db: 数据库会话

        Returns:
            BytesIO: Excel文件流
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "用户数据"

        # 表头定义
        headers = [
            ("账号", 15),
            ("姓名", 12),
            ("身份类型", 12),
            ("部门名称", 20),
            ("岗位名称", 15),
            ("邮箱", 25),
            ("手机号", 15),
            ("学号/工号", 15),
            ("入学/入职年份", 15),
            ("年级", 10),
            ("专业", 20),
            ("职称", 15),
            ("研究方向", 20),
            ("办公室", 15),
            ("紧急联系人", 12),
            ("紧急联系电话", 15)
        ]

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 添加必填标记行
        required_marks = [
            "是*", "是*", "是*", "是*",
            "否", "否", "否", "否",
            "否", "否", "否", "否",
            "否", "否", "否", "否"
        ]
        for col_idx, mark in enumerate(required_marks, 1):
            cell = ws.cell(row=2, column=col_idx, value=mark)
            cell.alignment = Alignment(horizontal='center')
            if "是" in mark:
                cell.fill = required_fill

        # 添加示例数据
        example_data = [
            "zhangsan", "张三", "student", "计算机学院",
            "", "zhangsan@example.com", "13800138000", "2024001",
            "2024", "大一", "软件工程", "", "", "", "", ""
        ]
        for col_idx, value in enumerate(example_data, 1):
            ws.cell(row=3, column=col_idx, value=value)

        # 添加说明Sheet
        ws_instructions = wb.create_sheet("填写说明")
        instructions = [
            "【批量用户导入模板填写说明】",
            "",
            "1. 必填字段：账号、姓名、身份类型、部门名称",
            "2. 身份类型只能填写：student（学生）或 teacher（老师）",
            "3. 部门名称必须是系统中已存在的部门",
            "4. 岗位名称必须是系统中已存在的岗位（学生可不填）",
            "5. 账号长度3-50字符，姓名长度2-50字符",
            "6. 邮箱格式：xxx@xxx.xxx",
            "7. 手机号格式：11位数字，1开头",
            "8. 默认密码：123456（可在导入时修改）",
            "",
            "【身份类型说明】",
            "student - 学生：需要填写学号、年级、专业等信息",
            "teacher - 老师：需要填写工号、职称、研究方向等信息",
        ]
        for row_idx, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row_idx, column=1, value=text)

        # 获取系统中的部门和岗位列表
        departments = db.query(Department).filter(
            Department.tenant_id == tenant_id
        ).all()

        positions = db.query(Position).filter(
            Position.tenant_id == tenant_id
        ).all()

        # 添加部门列表Sheet
        ws_depts = wb.create_sheet("部门列表")
        ws_depts.cell(row=1, column=1, value="部门名称")
        ws_depts.cell(row=1, column=2, value="部门类型")
        for row_idx, dept in enumerate(departments, 2):
            ws_depts.cell(row=row_idx, column=1, value=dept.name)
            ws_depts.cell(row=row_idx, column=2, value=dept.type)

        # 添加岗位列表Sheet
        ws_positions = wb.create_sheet("岗位列表")
        ws_positions.cell(row=1, column=1, value="岗位名称")
        for row_idx, pos in enumerate(positions, 2):
            ws_positions.cell(row=row_idx, column=1, value=pos.name)

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def parse_excel(file_content: bytes) -> List[Dict[str, Any]]:
        """
        解析Excel文件

        Args:
            file_content: Excel文件内容

        Returns:
            List[Dict]: 解析后的行数据列表
        """
        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active

            rows = []
            for row_idx in range(2, ws.max_row + 1):  # 跳过表头
                row_data = {}
                has_data = False

                for col_letter, field_name in BatchImportService.COLUMN_MAP.items():
                    col_idx = ord(col_letter) - ord('A') + 1
                    cell_value = ws.cell(row=row_idx, column=col_idx).value

                    if cell_value is not None:
                        has_data = True
                        row_data[field_name] = str(cell_value).strip() if isinstance(cell_value, str) else cell_value

                if has_data:
                    row_data['row_number'] = row_idx
                    rows.append(row_data)

            return rows

        except Exception as e:
            logger.error(f"解析Excel文件失败: {str(e)}")
            raise ValidationError(f"Excel文件解析失败: {str(e)}")

    @staticmethod
    def validate_row(
        row_data: Dict[str, Any],
        tenant_id: int,
        db: Session,
        existing_accounts: set
    ) -> Tuple[Optional[BatchImportUserRow], Optional[str]]:
        """
        验证单行数据

        Args:
            row_data: 行数据
            tenant_id: 租户ID
            db: 数据库会话
            existing_accounts: 已存在的账号集合

        Returns:
            Tuple[验证后的数据, 错误信息]
        """
        errors = []

        # 检查必填字段
        required_fields = ['account', 'name', 'identity_type', 'department_name']
        for field in required_fields:
            if not row_data.get(field):
                errors.append(f"缺少必填字段: {field}")

        # 检查账号唯一性
        account = row_data.get('account', '')
        if account:
            if account in existing_accounts:
                errors.append(f"账号 '{account}' 在导入文件中重复")
            else:
                # 检查数据库中是否已存在
                existing_user = db.query(User).filter(
                    User.tenant_id == tenant_id,
                    User.account == account
                ).first()
                if existing_user:
                    errors.append(f"账号 '{account}' 在系统中已存在")
                else:
                    existing_accounts.add(account)

        # 检查部门是否存在
        dept_name = row_data.get('department_name', '')
        if dept_name:
            dept = db.query(Department).filter(
                Department.tenant_id == tenant_id,
                Department.name == dept_name
            ).first()
            if not dept:
                errors.append(f"部门 '{dept_name}' 不存在")

        # 检查岗位是否存在（如果填写了）
        pos_name = row_data.get('position_name', '')
        if pos_name:
            pos = db.query(Position).filter(
                Position.tenant_id == tenant_id,
                Position.name == pos_name
            ).first()
            if not pos:
                errors.append(f"岗位 '{pos_name}' 不存在")

        # 检查邮箱格式
        email = row_data.get('email', '')
        if email and '@' not in email:
            errors.append("邮箱格式不正确")

        # 检查手机号格式
        phone = row_data.get('phone', '')
        if phone and (len(str(phone)) != 11 or not str(phone).startswith('1')):
            errors.append("手机号格式不正确")

        if errors:
            return None, "; ".join(errors)

        # 构建验证后的数据
        try:
            validated_data = BatchImportUserRow(**row_data)
            return validated_data, None
        except Exception as e:
            return None, str(e)

    @staticmethod
    def save_import_log(
        filename: str,
        total_rows: int,
        success_count: int,
        failed_count: int,
        default_password: str,
        results: List[BatchImportRowResult],
        operator_user_id: int,
        tenant_id: int,
        db: Session
    ) -> BatchImportLog:
        """
        保存导入日志

        Args:
            filename: 文件名
            total_rows: 总行数
            success_count: 成功数量
            failed_count: 失败数量
            default_password: 默认密码
            results: 处理结果列表
            operator_user_id: 操作人ID
            tenant_id: 租户ID
            db: 数据库会话

        Returns:
            BatchImportLog: 导入日志记录
        """
        # 收集错误详情
        error_details = []
        for result in results:
            if not result.success:
                error_details.append({
                    "row": result.row_number,
                    "account": result.account,
                    "name": result.name,
                    "error": result.error_message
                })

        import_log = BatchImportLog(
            tenant_id=tenant_id,
            filename=filename,
            total_rows=total_rows,
            success_count=success_count,
            failed_count=failed_count,
            default_password=default_password,
            error_details=json.dumps(error_details, ensure_ascii=False) if error_details else None,
            created_by=operator_user_id
        )
        db.add(import_log)
        db.flush()

        return import_log

    @staticmethod
    def import_users(
        file_content: bytes,
        filename: str,
        tenant_id: int,
        default_password: str,
        operator_user_id: int,
        db: Session,
        default_department_id: Optional[int] = None
    ) -> BatchImportResponse:
        """
        批量导入用户

        Args:
            file_content: Excel文件内容
            filename: 文件名
            tenant_id: 租户ID
            default_password: 默认密码
            operator_user_id: 操作人ID
            db: 数据库会话
            default_department_id: 默认部门ID（可选）

        Returns:
            BatchImportResponse: 导入结果
        """
        # 解析Excel
        rows = BatchImportService.parse_excel(file_content)

        if not rows:
            raise ValidationError("Excel文件中没有数据")

        # 获取默认角色
        student_role = db.query(Role).filter(
            Role.tenant_id == tenant_id,
            Role.name == "学生"
        ).first()

        teacher_role = db.query(Role).filter(
            Role.tenant_id == tenant_id,
            Role.name == "老师"
        ).first()

        # 处理结果
        results = []
        success_count = 0
        failed_count = 0
        existing_accounts = set()
        hashed_password = hash_password(default_password)

        for row_data in rows:
            row_number = row_data.get('row_number', 0)

            # 验证数据
            validated_data, error_msg = BatchImportService.validate_row(
                row_data, tenant_id, db, existing_accounts
            )

            if error_msg:
                results.append(BatchImportRowResult(
                    row_number=row_number,
                    success=False,
                    account=row_data.get('account'),
                    name=row_data.get('name'),
                    error_message=error_msg
                ))
                failed_count += 1
                continue

            try:
                # 获取部门ID
                dept = None
                if validated_data.department_name:
                    dept = db.query(Department).filter(
                        Department.tenant_id == tenant_id,
                        Department.name == validated_data.department_name
                    ).first()
                
                # 如果没有找到部门，使用默认部门
                if not dept and default_department_id:
                    dept = db.query(Department).filter(
                        Department.tenant_id == tenant_id,
                        Department.id == default_department_id
                    ).first()
                
                # 如果仍然没有部门，跳过该用户
                if not dept:
                    results.append(BatchImportRowResult(
                        row_number=row_number,
                        success=False,
                        account=validated_data.account,
                        name=validated_data.name,
                        error_message="未找到对应的部门，请在Excel中指定部门或设置默认部门"
                    ))
                    failed_count += 1
                    continue

                # 获取岗位ID（如果指定了）
                pos = None
                if validated_data.position_name:
                    pos = db.query(Position).filter(
                        Position.tenant_id == tenant_id,
                        Position.name == validated_data.position_name
                    ).first()

                # 创建用户
                new_user = User(
                    tenant_id=tenant_id,
                    account=validated_data.account,
                    name=validated_data.name,
                    password_hash=hashed_password,
                    email=validated_data.email,
                    phone=validated_data.phone,
                    department_id=dept.id,
                    is_active=True
                )
                db.add(new_user)
                db.flush()

                # 创建用户扩展信息
                profile = UserProfile(
                    user_id=new_user.id,
                    identity_no=validated_data.identity_no,
                    identity_type=validated_data.identity_type,
                    entry_year=validated_data.entry_year,
                    grade=validated_data.grade,
                    major=validated_data.major,
                    title=validated_data.title,
                    research_area=validated_data.research_area,
                    office=validated_data.office,
                    emergency_contact=validated_data.emergency_contact,
                    emergency_phone=validated_data.emergency_phone
                )
                db.add(profile)

                # 创建用户-部门关联
                user_dept = UserDepartment(
                    tenant_id=tenant_id,
                    user_id=new_user.id,
                    department_id=dept.id,
                    is_primary=True
                )
                db.add(user_dept)

                # 创建用户-部门-岗位关联（如果指定了岗位）
                if pos:
                    user_dept_post = UserDepartmentPost(
                        tenant_id=tenant_id,
                        user_id=new_user.id,
                        department_id=dept.id,
                        post_id=pos.id
                    )
                    db.add(user_dept_post)

                # 分配角色
                role = student_role if validated_data.identity_type == 'student' else teacher_role
                if role:
                    user_role = UserRole(
                        tenant_id=tenant_id,
                        user_id=new_user.id,
                        role_id=role.id
                    )
                    db.add(user_role)

                # 提交当前用户
                db.commit()

                results.append(BatchImportRowResult(
                    row_number=row_number,
                    success=True,
                    account=validated_data.account,
                    name=validated_data.name,
                    user_id=new_user.id
                ))
                success_count += 1

            except Exception as e:
                db.rollback()
                logger.error(f"创建用户失败 (行{row_number}): {str(e)}")
                results.append(BatchImportRowResult(
                    row_number=row_number,
                    success=False,
                    account=validated_data.account,
                    name=validated_data.name,
                    error_message=str(e)
                ))
                failed_count += 1
                continue

        # 保存导入日志
        BatchImportService.save_import_log(
            filename=filename,
            total_rows=len(rows),
            success_count=success_count,
            failed_count=failed_count,
            default_password=default_password,
            results=results,
            operator_user_id=operator_user_id,
            tenant_id=tenant_id,
            db=db
        )
        db.commit()

        return BatchImportResponse(
            total_rows=len(rows),
            success_count=success_count,
            failed_count=failed_count,
            results=results,
            default_password=default_password
        )

    @staticmethod
    async def preview_import(
        file_content: bytes,
        tenant_id: int,
        db: Session
    ) -> ImportPreviewResponse:
        """
        预览导入数据

        Args:
            file_content: Excel文件内容
            tenant_id: 租户ID
            db: 数据库会话

        Returns:
            ImportPreviewResponse: 预览数据
        """
        # 解析Excel
        rows = BatchImportService.parse_excel(file_content)

        if not rows:
            raise ValidationError("Excel文件中没有数据")

        # 验证每一行
        preview_rows = []
        valid_count = 0
        invalid_count = 0
        existing_accounts = set()

        for row_data in rows:
            row_number = row_data.get('row_number', 0)
            validated_data, error_msg = BatchImportService.validate_row(
                row_data, tenant_id, db, existing_accounts
            )

            is_valid = error_msg is None
            if is_valid:
                valid_count += 1
            else:
                invalid_count += 1

            preview_row = ImportPreviewRow(
                row_index=row_number,
                account=row_data.get('account', ''),
                name=row_data.get('name', ''),
                email=row_data.get('email'),
                phone=str(row_data.get('phone', '')) if row_data.get('phone') else None,
                department_name=row_data.get('department_name', ''),
                position_name=row_data.get('position_name'),
                role=row_data.get('identity_type'),
                is_valid=is_valid,
                error_message=error_msg
            )
            preview_rows.append(preview_row)

        # 生成preview_key
        preview_key = f"import_preview:{tenant_id}:{uuid.uuid4().hex}"

        # 存储预览数据到Redis（30分钟过期）
        preview_data = {
            "rows": [row.model_dump() for row in preview_rows],
            "file_content": file_content.hex()
        }
        await redis_client.set(preview_key, preview_data, expire=1800)

        logger.info(f"生成导入预览: key={preview_key}, total={len(rows)}, valid={valid_count}, invalid={invalid_count}")

        return ImportPreviewResponse(
            preview_key=preview_key,
            total_rows=len(rows),
            valid_rows=valid_count,
            invalid_rows=invalid_count,
            rows=preview_rows
        )

    @staticmethod
    async def confirm_import(
        preview_key: str,
        selected_rows: Optional[List[int]],
        tenant_id: int,
        operator_user_id: int,
        db: Session
    ) -> BatchImportResponse:
        """
        确认导入数据

        Args:
            preview_key: 预览key
            selected_rows: 选中的行索引列表，为空则导入全部有效行
            tenant_id: 租户ID
            operator_user_id: 操作人ID
            db: 数据库会话

        Returns:
            BatchImportResponse: 导入结果
        """
        # 从Redis获取预览数据
        preview_data = await redis_client.get(preview_key)
        if not preview_data:
            raise ValidationError("预览数据已过期，请重新上传文件")

        rows_data = preview_data.get("rows", [])
        if not rows_data:
            raise ValidationError("预览数据为空")

        # 确定要导入的行
        if selected_rows:
            # 导入选中的行
            rows_to_import = [r for r in rows_data if r.get("row_index") in selected_rows and r.get("is_valid")]
        else:
            # 导入全部有效行
            rows_to_import = [r for r in rows_data if r.get("is_valid")]

        if not rows_to_import:
            raise ValidationError("没有可导入的有效数据")

        # 解析文件内容重新获取原始数据
        file_content = bytes.fromhex(preview_data.get("file_content", ""))
        original_rows = BatchImportService.parse_excel(file_content)

        # 构建行号到原始数据的映射
        row_map = {r.get("row_number"): r for r in original_rows}

        # 获取默认角色
        student_role = db.query(Role).filter(
            Role.tenant_id == tenant_id,
            Role.name == "学生"
        ).first()

        teacher_role = db.query(Role).filter(
            Role.tenant_id == tenant_id,
            Role.name == "老师"
        ).first()

        # 执行导入
        results = []
        success_count = 0
        failed_count = 0
        existing_accounts = set()
        hashed_password = hash_password("123456")

        for import_row in rows_to_import:
            row_index = import_row.get("row_index")
            row_data = row_map.get(row_index)

            if not row_data:
                continue

            # 验证数据
            validated_data, error_msg = BatchImportService.validate_row(
                row_data, tenant_id, db, existing_accounts
            )

            if error_msg:
                results.append(BatchImportRowResult(
                    row_number=row_index,
                    success=False,
                    account=row_data.get('account'),
                    name=row_data.get('name'),
                    error_message=error_msg
                ))
                failed_count += 1
                continue

            try:
                # 获取部门ID
                dept = db.query(Department).filter(
                    Department.tenant_id == tenant_id,
                    Department.name == validated_data.department_name
                ).first()

                # 获取岗位ID（如果指定了）
                pos = None
                if validated_data.position_name:
                    pos = db.query(Position).filter(
                        Position.tenant_id == tenant_id,
                        Position.name == validated_data.position_name
                    ).first()

                # 创建用户
                new_user = User(
                    tenant_id=tenant_id,
                    account=validated_data.account,
                    name=validated_data.name,
                    password_hash=hashed_password,
                    email=validated_data.email,
                    phone=validated_data.phone,
                    department_id=dept.id,
                    is_active=True
                )
                db.add(new_user)
                db.flush()

                # 创建用户扩展信息
                profile = UserProfile(
                    user_id=new_user.id,
                    identity_no=validated_data.identity_no,
                    identity_type=validated_data.identity_type,
                    entry_year=validated_data.entry_year,
                    grade=validated_data.grade,
                    major=validated_data.major,
                    title=validated_data.title,
                    research_area=validated_data.research_area,
                    office=validated_data.office,
                    emergency_contact=validated_data.emergency_contact,
                    emergency_phone=validated_data.emergency_phone
                )
                db.add(profile)

                # 创建用户-部门关联
                user_dept = UserDepartment(
                    tenant_id=tenant_id,
                    user_id=new_user.id,
                    department_id=dept.id,
                    is_primary=True
                )
                db.add(user_dept)

                # 创建用户-部门-岗位关联（如果指定了岗位）
                if pos:
                    user_dept_post = UserDepartmentPost(
                        tenant_id=tenant_id,
                        user_id=new_user.id,
                        department_id=dept.id,
                        post_id=pos.id
                    )
                    db.add(user_dept_post)

                # 分配角色
                role = student_role if validated_data.identity_type == 'student' else teacher_role
                if role:
                    user_role = UserRole(
                        tenant_id=tenant_id,
                        user_id=new_user.id,
                        role_id=role.id
                    )
                    db.add(user_role)

                # 提交当前用户
                db.commit()

                results.append(BatchImportRowResult(
                    row_number=row_index,
                    success=True,
                    account=validated_data.account,
                    name=validated_data.name,
                    user_id=new_user.id
                ))
                success_count += 1

            except Exception as e:
                db.rollback()
                logger.error(f"创建用户失败 (行{row_index}): {str(e)}")
                results.append(BatchImportRowResult(
                    row_number=row_index,
                    success=False,
                    account=validated_data.account,
                    name=validated_data.name,
                    error_message=str(e)
                ))
                failed_count += 1
                continue

        # 保存导入日志
        BatchImportService.save_import_log(
            filename="preview_import",
            total_rows=len(rows_to_import),
            success_count=success_count,
            failed_count=failed_count,
            default_password="123456",
            results=results,
            operator_user_id=operator_user_id,
            tenant_id=tenant_id,
            db=db
        )
        db.commit()

        # 删除预览数据
        await redis_client.delete(preview_key)

        logger.info(f"确认导入完成: key={preview_key}, success={success_count}, failed={failed_count}")

        return BatchImportResponse(
            total_rows=len(rows_to_import),
            success_count=success_count,
            failed_count=failed_count,
            results=results,
            default_password="123456"
        )
