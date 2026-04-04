"""
提交相关API端点
"""
from fastapi import APIRouter, Depends, Query, Path, Body, UploadFile, File, Request
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime
from app.core.database import get_db
from app.api.deps import get_current_user, get_current_tenant_id
from app.schemas.submission_schemas import *
from app.services.submission_service import SubmissionService
from app.services.draft_service import DraftService
from app.services.attachment_service import AttachmentService
from app.services.export_service import ExportService
from app.core.response import success_response, error_response
from app.core.exceptions import ValidationError, BusinessError, NotFoundError
from app.models.user import User
from app.models.form import Submission
from app.utils.audit import audit_log
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


# ========== 提交管理 ==========

@router.post("", summary="创建提交")
@audit_log(action="create_submission", resource_type="submission", record_after=True)
async def create_submission(
        request: "SubmissionCreateRequest" = Body(...),
        req: Request = None,
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """创建提交"""
    try:
        logger.info(f"创建提交请求: form_id={request.form_id}, user_id={current_user.id}, data_keys={list(request.data.keys())}")
        
        # 获取IP和设备信息
        ip_address = req.client.host if req else None
        device_info = {
            "user_agent": req.headers.get("user-agent") if req else None
        }

        submission = SubmissionService.create_submission(
            request=request,
            tenant_id=tenant_id,
            user_id=current_user.id,
            ip_address=ip_address,
            device_info=device_info,
            db=db
        )

        return success_response(
            data=SubmissionResponse.from_orm(submission).dict(),
            message="提交成功"
        )
    except ValidationError as e:
        logger.error(f"Validation error in create_submission: {str(e)}")
        return error_response(str(e), 4001)
    except BusinessError as e:
        logger.error(f"Business error in create_submission: {str(e)}")
        return error_response(str(e), 4002)
    except Exception as e:
        logger.error(f"Create submission error: {e}", exc_info=True)
        return error_response("提交失败", 5001)


@router.put("/{submission_id}", summary="更新提交")
@audit_log(action="update_submission", resource_type="submission", record_before=True, record_after=True)
async def update_submission(
        submission_id: int = Path(...),
        request: "SubmissionUpdateRequest" = Body(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """更新提交"""
    try:
        submission = SubmissionService.update_submission(
            submission_id=submission_id,
            request=request,
            tenant_id=tenant_id,
            user_id=current_user.id,
            db=db
        )

        return success_response(
            data=SubmissionResponse.from_orm(submission).dict(),
            message="更新成功"
        )
    except ValidationError as e:
        return error_response(str(e), 4001)
    except BusinessError as e:
        return error_response(str(e), 4002)
    except Exception as e:
        logger.error(f"Update submission error: {e}")
        return error_response("更新失败", 5001)


@router.delete("/{submission_id}", summary="删除提交")
@audit_log(action="delete_submission", resource_type="submission", record_before=True)
async def delete_submission(
        submission_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """删除提交"""
    try:
        SubmissionService.delete_submission(
            submission_id=submission_id,
            tenant_id=tenant_id,
            user_id=current_user.id,
            db=db
        )

        return success_response(message="删除成功")
    except BusinessError as e:
        return error_response(str(e), 4002)
    except Exception as e:
        logger.error(f"Delete submission error: {e}")
        return error_response("删除失败", 5001)


@router.get("/{submission_id}", summary="获取提交详情")
async def get_submission(
        submission_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """获取提交详情"""
    try:
        from app.models.form import Form, FormVersion
        
        submission = SubmissionService.get_submission_by_id(
            submission_id=submission_id,
            tenant_id=tenant_id,
            db=db
        )

        # 获取表单信息
        form = db.query(Form).filter(Form.id == submission.form_id).first()
        form_name = form.name if form else "未知表单"

        # 获取附件
        attachments = AttachmentService.list_attachments(
            owner_type="submission",
            owner_id=submission_id,
            tenant_id=tenant_id,
            db=db
        )

        response_data = SubmissionDetailResponse.from_orm(submission).dict()
        response_data["form_name"] = form_name
        
        # 获取提交人姓名
        if submission.submitter_user_id:
            submitter = db.query(User).filter(User.id == submission.submitter_user_id).first()
            response_data["submitter_name"] = submitter.name if submitter else "未知用户"
        else:
            response_data["submitter_name"] = "匿名用户"
        
        # 为附件生成下载URL
        attachments_with_url = []
        for att in attachments:
            att_dict = AttachmentResponse.from_orm(att).dict()
            att_dict["download_url"] = AttachmentService.get_download_url(
                att.id, tenant_id, db=db
            )
            attachments_with_url.append(att_dict)
        response_data["attachments"] = attachments_with_url
        
        form_version = db.query(FormVersion).filter(
            FormVersion.id == submission.form_version_id
        ).first()

        if form_version and form_version.schema_json:
            schema = form_version.schema_json
            if isinstance(schema, dict) and "fields" in schema:
                if not response_data.get("snapshot_json"):
                    response_data["snapshot_json"] = {}
                
                snapshot = response_data["snapshot_json"]
                
                if not snapshot.get("field_labels"):
                    field_labels = {}
                    for field in schema["fields"]:
                        if isinstance(field, dict) and "id" in field:
                            field_labels[field["id"]] = field.get("label", field["id"])
                    snapshot["field_labels"] = field_labels
                
                if not snapshot.get("field_types"):
                    field_types = {}
                    for field in schema["fields"]:
                        if isinstance(field, dict) and "id" in field:
                            field_types[field["id"]] = field.get("type", "text")
                    snapshot["field_types"] = field_types
                
                if not snapshot.get("field_options"):
                    field_options = {}
                    for field in schema["fields"]:
                        if isinstance(field, dict) and "id" in field:
                            if field.get("type") in ["select", "radio", "checkbox"]:
                                options = field.get("props", {}).get("options", [])
                                field_options[field["id"]] = options
                    snapshot["field_options"] = field_options
        
        process_instance_id, process_state = SubmissionService.get_process_overview(
            submission_id=submission_id,
            tenant_id=tenant_id,
            db=db,
        )
        response_data["process_instance_id"] = process_instance_id
        response_data["process_state"] = process_state

        return success_response(data=response_data)
    except NotFoundError as e:
        return error_response(str(e), 4041)
    except Exception as e:
        logger.error(f"Get submission error: {e}")
        return error_response("查询失败", 5001)


@router.get("", summary="查询提交列表")
async def list_submissions(
        page: int = Query(1, ge=1),
        page_size: int = Query(20, ge=1, le=100),
        form_id: Optional[int] = Query(None),
        status: Optional[str] = Query(None),
        submitter_user_id: Optional[int] = Query(None),
        keyword: Optional[str] = Query(None),
        date_from: Optional[datetime] = Query(None),
        date_to: Optional[datetime] = Query(None),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """查询提交列表"""
    try:
        from app.models.form import Form
        from app.models.user import User
        from app.models.workflow import ProcessInstance, Task

        request = SubmissionQueryRequest(
            page=page,
            page_size=page_size,
            form_id=form_id,
            status=status,
            submitter_user_id=submitter_user_id,
            keyword=keyword,
            date_from=date_from,
            date_to=date_to
        )

        submissions, total = SubmissionService.list_submissions(
            request=request,
            tenant_id=tenant_id,
            user_id=current_user.id,
            db=db
        )

        # 批量查询关联的流程实例
        submission_ids = [s.id for s in submissions]
        process_instances = {}
        if submission_ids:
            processes = db.query(ProcessInstance).filter(
                ProcessInstance.submission_id.in_(submission_ids)
            ).all()
            for p in processes:
                process_instances[p.submission_id] = p

        # 批量查询当前进行中的任务（获取 SLA 信息）
        process_instance_ids = [p.id for p in process_instances.values()]
        current_tasks = {}
        now = datetime.utcnow()
        if process_instance_ids:
            tasks = db.query(Task).filter(
                Task.process_instance_id.in_(process_instance_ids),
                Task.status.in_(["open", "claimed"])
            ).all()
            for task in tasks:
                # 每个流程实例只取第一个未完成的任务
                if task.process_instance_id not in current_tasks:
                    current_tasks[task.process_instance_id] = task

        items = []
        for s in submissions:
            item_dict = SubmissionResponse.from_orm(s).dict()

            # 获取表单名称
            form = db.query(Form).filter(Form.id == s.form_id).first()
            item_dict["form_name"] = form.name if form else "未知表单"

            # 获取提交人名称
            if s.submitter_user_id:
                user = db.query(User).filter(User.id == s.submitter_user_id).first()
                item_dict["submitter_name"] = user.name if user else "未知用户"
            else:
                item_dict["submitter_name"] = "匿名用户"

            # 获取流程实例信息
            process = process_instances.get(s.id)
            if process:
                item_dict["process_instance_id"] = process.id
                item_dict["process_state"] = process.state
                item_dict["flow_definition_id"] = process.flow_definition_id

                # 获取当前任务的 SLA 信息
                current_task = current_tasks.get(process.id)
                if current_task:
                    item_dict["due_at"] = current_task.due_at.isoformat() if current_task.due_at else None
                    item_dict["is_overdue"] = bool(current_task.due_at and now > current_task.due_at)
                else:
                    item_dict["due_at"] = None
                    item_dict["is_overdue"] = False
            else:
                item_dict["process_instance_id"] = None
                item_dict["process_state"] = None
                item_dict["flow_definition_id"] = None
                item_dict["due_at"] = None
                item_dict["is_overdue"] = False

            items.append(item_dict)

        response = SubmissionListResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=(total + page_size - 1) // page_size
        )

        return success_response(data=response.dict())
    except Exception as e:
        logger.error(f"List submissions error: {e}")
        return error_response("查询失败", 5001)


@router.get("/forms/{form_id}/latest", summary="获取用户最新提交")
async def get_latest_submission(
        form_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """获取当前用户对指定表单的最新提交记录"""
    try:
        submission = SubmissionService.get_latest_submission_by_user(
            form_id=form_id,
            user_id=current_user.id,
            tenant_id=tenant_id,
            db=db
        )

        if not submission:
            return success_response(data=None, message="无历史提交")

        return success_response(data=SubmissionResponse.from_orm(submission).dict())
    except Exception as e:
        logger.error(f"Get latest submission error: {e}")
        return error_response("查询失败", 5001)


# ========== 草稿管理 ==========

@router.post("/drafts", summary="保存草稿")
async def save_draft(
        request: "DraftSaveRequest" = Body(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """保存草稿"""
    try:
        draft = DraftService.save_draft(
            request=request,
            tenant_id=tenant_id,
            user_id=current_user.get("user_id"),
            session_id=None,
            db=db
        )

        return success_response(
            data=DraftResponse.from_orm(draft).dict(),
            message="草稿已保存"
        )
    except Exception as e:
        logger.error(f"Save draft error: {e}")
        return error_response("保存失败", 5001)


@router.get("/drafts/{form_id}", summary="获取草稿")
async def get_draft(
        form_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """获取草稿"""
    try:
        draft = DraftService.get_draft(
            form_id=form_id,
            tenant_id=tenant_id,
            user_id=current_user.get("user_id"),
            session_id=None,
            db=db
        )

        if not draft:
            return success_response(data=None, message="无草稿")

        return success_response(data=DraftResponse.from_orm(draft).dict())
    except Exception as e:
        logger.error(f"Get draft error: {e}")
        return error_response("查询失败", 5001)


@router.delete("/drafts/{draft_id}", summary="删除草稿")
async def delete_draft(
        draft_id: int = Path(...),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """删除草稿"""
    try:
        DraftService.delete_draft(draft_id, tenant_id, db)
        return success_response(message="删除成功")
    except Exception as e:
        logger.error(f"Delete draft error: {e}")
        return error_response("删除失败", 5001)


# ========== 统计分析 ==========

@router.get("/statistics/{form_id}", summary="提交统计")
async def get_statistics(
        form_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """提交统计"""
    try:
        stats = SubmissionService.get_submission_statistics(
            form_id=form_id,
            tenant_id=tenant_id,
            db=db
        )

        return success_response(data=stats)
    except Exception as e:
        logger.error(f"Get statistics error: {e}")
        return error_response("查询失败", 5001)


# ========== 数据导出 ==========

@router.post("/export", summary="导出数据")
async def export_submissions(
        request: "ExportRequest" = Body(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """导出数据"""
    try:
        result = ExportService.export_submissions(
            request=request,
            tenant_id=tenant_id,
            user_id=current_user.id,
            db=db
        )

        return success_response(data=result)
    except Exception as e:
        logger.error(f"Export submissions error: {e}")
        return error_response("导出失败", 5001)


@router.get("/export/{task_id}", summary="查询导出任务")
async def get_export_task(
        task_id: str = Path(...),
        tenant_id: int = Depends(get_current_tenant_id)
):
    """查询导出任务状态"""
    try:
        status = ExportService.get_export_task_status(task_id, tenant_id)
        return success_response(data=status)
    except NotFoundError as e:
        return error_response(str(e), 4041)
    except Exception as e:
        logger.error(f"Get export task error: {e}")
        return error_response("查询失败", 5001)


# ========== 审批流程 ==========

@router.post("/{submission_id}/start-approval", summary="发起审批")
@audit_log(action="start_approval", resource_type="submission", record_before=True, record_after=True)
async def start_approval(
        submission_id: int = Path(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """手动发起审批流程（用于 pending_approval 状态的提交）"""
    try:
        submission = SubmissionService.start_approval(
            submission_id=submission_id,
            tenant_id=tenant_id,
            user_id=current_user.id,
            db=db
        )

        return success_response(
            data=SubmissionResponse.from_orm(submission).dict(),
            message="审批流程已发起"
        )
    except NotFoundError as e:
        return error_response(str(e), 4041)
    except BusinessError as e:
        return error_response(str(e), 4002)
    except Exception as e:
        logger.error(f"Start approval error: {e}", exc_info=True)
        return error_response("发起审批失败", 5001)


def _convert_field_value(field: dict, raw_value, label_to_value: dict) -> any:
    """根据字段类型转换 Excel 单元格值。"""
    from datetime import datetime, date, time as dt_time

    ftype = field.get("type", "text")
    fid = field.get("id", "")
    val = raw_value

    if ftype in ("select", "radio"):
        mapping = label_to_value.get(fid, {})
        return mapping.get(str(val), str(val))

    if ftype == "checkbox":
        mapping = label_to_value.get(fid, {})
        parts = [s.strip() for s in str(val).split(",") if s.strip()]
        return [mapping.get(p, p) for p in parts]

    if ftype == "switch":
        s = str(val).strip().lower()
        return s in ("是", "true", "1", "yes")

    if ftype in ("date", "datetime"):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, date):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip().replace("/", "-")
        return s

    if ftype in ("date_range", "date-range"):
        import re
        
        # 处理单个 datetime 对象（Excel 中可能是单个日期）
        if isinstance(val, datetime):
            d = val.strftime("%Y-%m-%d")
            return {"start": d, "end": d}
        
        # 处理列表/tuple类型的日期范围
        if hasattr(val, '__iter__') and not isinstance(val, str):
            try:
                items = list(val)
                if len(items) >= 2:
                    start_dt = items[0]
                    end_dt = items[1]
                    if hasattr(start_dt, 'strftime'):
                        return {"start": start_dt.strftime("%Y-%m-%d"), "end": end_dt.strftime("%Y-%m-%d")}
            except (TypeError, IndexError, KeyError):
                pass
        
        # 解析字符串中的日期
        s = str(val).strip().replace("/", "-").replace("\\", "-") if val else ""
        
        # 使用正则表达式提取两个日期（更可靠的方法）
        date_pattern = r'\d{4}-\d{2}-\d{2}'
        dates = re.findall(date_pattern, s)
        if len(dates) == 2:
            return {"start": dates[0], "end": dates[1]}
        
        # 如果正则只找到一个日期，返回单个日期
        if len(dates) == 1:
            return {"start": dates[0], "end": dates[0]}
        
        return s

    if ftype == "time":
        if isinstance(val, dt_time):
            return val.strftime("%H:%M")
        if isinstance(val, datetime):
            return val.strftime("%H:%M")
        return str(val).strip()

    if ftype in ("number", "rate"):
        try:
            return float(val)
        except (ValueError, TypeError):
            return val

    return str(val).strip()


@router.post("/batch-import", summary="批量导入表单数据（创建草稿）")
@audit_log(action="batch_import", resource_type="submission")
async def batch_import(
        form_id: int = Body(..., embed=True),
        file: UploadFile = File(...),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """解析 Excel 文件，为每行数据创建一条草稿提交。"""
    from openpyxl import load_workbook
    from io import BytesIO
    from app.models.form import Form, FormVersion
    from app.services.formula_service import FormulaService

    form = db.query(Form).filter(Form.id == form_id, Form.tenant_id == tenant_id).first()
    if not form:
        raise NotFoundError("表单不存在")

    # 权限校验：表单创建者或有填写/管理权限的用户
    from app.services.form_permission_service import FormPermissionService
    if form.owner_user_id != current_user.id:
        has_perm = FormPermissionService.has_permission(
            form_id=form_id,
            tenant_id=tenant_id,
            user_id=current_user.id,
            perm_type="fill",
            db=db,
        )
        if not has_perm:
            has_manage = FormPermissionService.has_permission(
                form_id=form_id,
                tenant_id=tenant_id,
                user_id=current_user.id,
                perm_type="manage",
                db=db,
            )
            if not has_manage:
                raise AuthorizationError("无权限导入此表单数据")

    version = db.query(FormVersion).filter(
        FormVersion.form_id == form_id,
        FormVersion.version > 0,
    ).order_by(FormVersion.version.desc()).first()
    if not version:
        raise NotFoundError("表单未发布")

    schema_json = version.schema_json
    fields = schema_json.get("fields", [])
    SKIP_TYPES = {"description", "divider", "calculated", "upload"}
    import_fields = [f for f in fields if f.get("type") not in SKIP_TYPES]

    # 预构建 label→value 映射
    label_to_value: dict[str, dict[str, str]] = {}
    for f in import_fields:
        fid = f.get("id")
        options = f.get("props", {}).get("options", [])
        if options:
            mapping = {}
            for opt in options:
                lbl = str(opt.get("label", opt.get("value", "")))
                val = str(opt.get("value", lbl))
                mapping[lbl] = val
                mapping[val] = val
            label_to_value[fid] = mapping

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    # 跳过表头(行1)、格式说明(行2)和示例数据(行3)，从行4开始读数据
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()

    if not rows:
        return error_response("Excel 文件中没有数据行", 4001)

    # 收集计算字段
    calc_fields = [f for f in fields if f.get("type") == "calculated"]

    created = []
    for row_idx, row in enumerate(rows, start=3):
        if all(cell is None for cell in row):
            continue
        data = {}
        for col_idx, field in enumerate(import_fields):
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    converted = _convert_field_value(field, val, label_to_value)
                    data[field["id"]] = converted
        if not data:
            continue

        # 计算字段自动求值
        for cf in calc_fields:
            formula = cf.get("props", {}).get("formula", "")
            if formula:
                try:
                    data[cf["id"]] = FormulaService.evaluate(formula, data)
                except Exception:
                    pass

        from app.services.submission_service import SubmissionService
        snapshot = SubmissionService.build_snapshot(version)
        
        submission = Submission(
            form_id=form_id,
            form_version_id=version.id,
            tenant_id=tenant_id,
            submitter_user_id=current_user.id,
            data_jsonb=data,
            snapshot_json=snapshot,
            status="draft",
        )
        db.add(submission)
        db.flush()
        created.append({"id": submission.id, "row": row_idx})

    db.commit()

    return success_response(
        data={"created": len(created), "items": created},
        message=f"成功导入 {len(created)} 条草稿",
    )


@router.post("/batch-submit", summary="批量提交审批")
@audit_log(action="batch_submit", resource_type="submission")
async def batch_submit(
        submission_ids: List[int] = Body(..., embed=True),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """将多条 draft 状态的提交批量发起审批流程。"""
    if not submission_ids:
        return error_response("请选择要提交的记录", 4001)

    results = []
    for sid in submission_ids:
        try:
            submission = SubmissionService.start_approval(
                submission_id=sid,
                tenant_id=tenant_id,
                user_id=current_user.id,
                db=db,
            )
            results.append({"id": sid, "success": True})
        except Exception as e:
            results.append({"id": sid, "success": False, "error": str(e)})

    db.commit()

    success_count = sum(1 for r in results if r["success"])
    return success_response(
        data={"total": len(results), "success": success_count, "items": results},
        message=f"批量提交完成：成功 {success_count}/{len(results)}",
    )


@router.post("/batch-approve", summary="批量通过（确认导入数据）")
@audit_log(action="batch_approve", resource_type="submission")
async def batch_approve(
        submission_ids: List[int] = Body(..., embed=True),
        current_user: User = Depends(get_current_user),
        tenant_id: int = Depends(get_current_tenant_id),
        db: Session = Depends(get_db)
):
    """将多条 draft 状态的提交标记为已通过并自动发起审批流程。"""
    from app.services.process_service import ProcessService
    from app.services.formula_service import FormulaService
    from app.models.form import FormVersion
    
    if not submission_ids:
        return error_response("请选择要通过的记录", 4001)

    results = []
    for sid in submission_ids:
        submission = db.query(Submission).filter(
            Submission.id == sid,
            Submission.tenant_id == tenant_id,
            Submission.submitter_user_id == current_user.id,
            Submission.status == "draft",
        ).first()
        if not submission:
            results.append({"id": sid, "success": False, "error": "记录不存在或状态不正确"})
            continue
        
        try:
            submission.status = "pending_approval"
            db.flush()
            
            form_version = db.query(FormVersion).filter(
                FormVersion.id == submission.form_version_id
            ).first()
            
            try:
                process = ProcessService.start_process(
                    form_id=submission.form_id,
                    form_version_id=submission.form_version_id,
                    submission_id=submission.id,
                    tenant_id=tenant_id,
                    db=db
                )
                submission.status = "submitted"
                logger.info(f"批量通过时自动创建流程实例成功: submission_id={sid}, process_id={process.id}")
            except BusinessError as e:
                # 流程创建失败，回滚 submission 状态
                logger.warning(f"批量通过时自动创建流程实例失败，已回滚状态: submission_id={sid}, error={e}")
                submission.status = "draft"
                db.flush()
                results.append({"id": sid, "success": False, "error": f"流程创建失败: {e}"})
                continue
            except Exception as e:
                # 其他异常也回滚状态
                logger.error(f"批量通过时创建流程实例异常，已回滚状态: submission_id={sid}, error={e}")
                submission.status = "draft"
                db.flush()
                results.append({"id": sid, "success": False, "error": f"流程创建异常: {e}"})
                continue
            
            results.append({"id": sid, "success": True})
        except Exception as e:
            logger.error(f"批量通过处理失败: submission_id={sid}, error={e}")
            results.append({"id": sid, "success": False, "error": str(e)})

    db.commit()
    success_count = sum(1 for r in results if r["success"])
    return success_response(
        data={"total": len(results), "success": success_count, "items": results},
        message=f"批量通过完成：成功 {success_count}/{len(results)}",
    )